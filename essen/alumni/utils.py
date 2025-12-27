import openpyxl
from .models import Alumnus

def import_alumni_from_xlsx(file_obj, stdout=None):
    """
    Import alumni from an openpyxl-compatible file object (path or stream).
    stdout: optional file-like object to write status messages to.
    Returns the count of imported alumni.
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        if stdout:
            stdout.write(f"Failed to load workbook: {e}\n")
        raise e

    count = 0
    for sheet_name in wb.sheetnames:
        # Try to parse year from sheet name
        try:
            year = int(sheet_name)
        except ValueError:
            if stdout:
                stdout.write(f"Skipping sheet '{sheet_name}' (not a year)\n")
            continue
        
        if stdout:
            stdout.write(f"Processing sheet '{sheet_name}' (Year: {year})...\n")
        
        ws = wb[sheet_name]
        
        # Iterate over rows, skipping the header
        rows = ws.iter_rows(min_row=2, values_only=True)
        
        for row in rows:
            # Ensure row has enough columns
            if not row or len(row) < 5:
                continue
            
            first_name = str(row[0]).strip() if row[0] else ""
            last_name = str(row[1]).strip() if row[1] else ""
            kerb = str(row[2]).strip() if row[2] else ""
            url = str(row[4]).strip() if row[4] else ""
            
            # Some rows might be empty or just comments
            if not first_name or not last_name:
                continue

            # Create or update
            defaults = {
                'kerb': kerb,
                'year': year,
                'url': url,
            }

            Alumnus.objects.update_or_create(
                first_name=first_name,
                last_name=last_name,
                defaults=defaults
            )
            count += 1
            
    return count
